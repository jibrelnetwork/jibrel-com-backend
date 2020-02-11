from operator import attrgetter

from django.http import HttpResponse
from django.utils.encoding import force_text
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    Side
)
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.utils import get_column_letter

side = Side(style=BORDER_THIN)

style_bold = NamedStyle(name='bold')
style_bold.font = Font(bold=True)

style_header = NamedStyle(name='header')
style_header.font = Font(bold=True)
style_header.border = Border(top=side, right=side, left=side, bottom=side)

style_content = NamedStyle(name='content')
style_content.border = Border(top=side, right=side, left=side, bottom=side)


def formatter_default(cell):
    return cell


def formatter_str(cell):
    return cell


def formatter_int(cell):
    cell.number_format = '0'
    return cell


def formatter_float(cell):
    cell.number_format = '0.00'


def formatter_decimal(cell):
    cell.number_format = '0.00'


def formatter_date(cell):
    cell.number_format = 'DD.MM.YYYY'
    cell.alignment = Alignment(horizontal="left")


def formatter_datetime(cell):
    cell.number_format = 'DD.MM.YYYY'
    cell.alignment = Alignment(horizontal="left")


def formatter_percent(cell):
    cell.style = 'Percent'


def formatter_currency(cell):
    cell.style = 'Currency'


class Column:
    number = 1
    letter = None
    title = None
    getter = None
    formatter = None
    width = 15

    def __init__(self, number, title, width, getter, formatter=None):
        self.number = number + 1
        self.letter = get_column_letter(self.number)
        self.title = force_text(title)
        self.width = width
        self.getter = getter if callable(getter) else attrgetter(getter)
        self.formatter = globals().get(f'formatter_{formatter}') or formatter_default

    def format(self, obj):
        value = self.getter(obj)
        if self.formatter is formatter_default:
            return value or ''
        return force_text(value)


def get_xlsx(mapper, data: list, filename: str = None, auto_filters: list = None, header: str = '') -> HttpResponse:
    filename = filename or 'data'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="%s.xlsx"' % filename or 'datasheet'

    wb = Workbook()
    ws = wb.active

    row_num = 0
    header_rows = header.strip('\n').split('\n')
    columns = [Column(number, *options) for number, options in enumerate(mapper)]

    for line in header_rows:
        row_num += 1
        c = ws.cell(row=row_num, column=1)
        c.value = force_text(line)
        c.style = style_bold

        ws.merge_cells('A{row_num}:{letter}{row_num}'.format(
            row_num=row_num,
            letter=get_column_letter(len(columns))
        ))

    # set column titles
    row_num += 1 + int(bool(header))
    for col in columns:
        c = ws.cell(row=row_num, column=col.number)
        c.value = col.title
        c.style = style_header
        ws.column_dimensions[col.letter].width = col.width

    # set data
    for obj in data:
        row_num += 1
        for col in columns:
            c = ws.cell(row=row_num, column=col.number)
            c.style = style_content
            col.formatter(c)  # type: ignore
            c.value = col.format(obj)

    if auto_filters:
        ws.auto_filter.ref = 'A2:%(letter)s%(length)s'.format(
            start=len(header_rows) + 1,
            letter=get_column_letter(len(columns)+1),
            length=len(header_rows) + len(data) + 1
        )
    wb.save(response)
    return response
